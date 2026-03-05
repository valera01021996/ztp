"""
Импорт данных из Excel в NetBox.

Использование:
    python import_to_netbox.py --file ../excel_templates/netbox_import.xlsx
    python import_to_netbox.py --file data.xlsx --dry-run   # без записи в NetBox
    python import_to_netbox.py --file data.xlsx --sheet Devices  # только один лист

Переменные окружения (.env):
    NETBOX_URL   = http://localhost:8000
    NETBOX_TOKEN = ваш_токен
"""

import os
import sys
import argparse
import logging
from typing import Optional

import pynetbox
from openpyxl import load_workbook
from dotenv import load_dotenv

load_dotenv(os.path.join(os.path.dirname(__file__), ".env"))

# ─── Логирование ─────────────────────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("netbox-import")


# ─── Вспомогательные функции ─────────────────────────────────────────────────

def cell_val(row, col_map: dict, key: str):
    """Безопасно достаёт значение ячейки по имени столбца."""
    idx = col_map.get(key)
    if idx is None:
        return None
    val = row[idx - 1].value
    if val is None:
        return None
    return str(val).strip() if isinstance(val, str) else val


def str_or_none(v) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


# Маппинг упрощённых названий → корректные значения NetBox API
IFACE_TYPE_MAP = {
    # Медь
    "100base-t":          "100base-tx",
    "1000base-t":         "1000base-t",
    "2.5gbase-t":         "2.5gbase-t",
    "5gbase-t":           "5gbase-t",
    "10gbase-t":          "10gbase-t",
    # SFP / SFP+
    "1g":                 "1000base-t",
    "1gbase-sx":          "1000base-x-sfp",
    "1gbase-lx":          "1000base-x-sfp",
    "10g":                "10gbase-x-sfpp",
    "10gbase-sr":         "10gbase-x-sfpp",
    "10gbase-lr":         "10gbase-x-sfpp",
    "10gbase-sfp+":       "10gbase-x-sfpp",
    "sfp+":               "10gbase-x-sfpp",
    # SFP28 / 25G
    "25g":                "25gbase-x-sfp28",
    "25gbase-sr":         "25gbase-x-sfp28",
    "25gbase-lr":         "25gbase-x-sfp28",
    # QSFP+ / 40G
    "40g":                "40gbase-x-qsfpp",
    "40gbase-sr4":        "40gbase-x-qsfpp",
    "qsfp+":              "40gbase-x-qsfpp",
    # QSFP28 / 100G  (100gbase-t не существует в NetBox — только copper 40G)
    "100g":               "100gbase-x-qsfp28",
    "100gbase-t":         "100gbase-x-qsfp28",   # частая ошибка — маппим на QSFP28
    "100gbase-sr4":       "100gbase-x-qsfp28",
    "100gbase-lr4":       "100gbase-x-qsfp28",
    "qsfp28":             "100gbase-x-qsfp28",
    # 400G
    "400g":               "400gbase-x-qsfpdd",
    "qsfp-dd":            "400gbase-x-qsfpdd",
    # Virtual / LAG
    "virtual":            "virtual",
    "lag":                "lag",
    "port-channel":       "lag",
    "ae":                 "lag",
    "bond":               "lag",
}


def normalize_iface_type(t: str) -> str:
    """Приводит тип интерфейса к формату NetBox. Неизвестные типы пропускаются as-is."""
    if not t:
        return "1000base-t"
    key = t.strip().lower()
    return IFACE_TYPE_MAP.get(key, key)


def status_val(v, default: str = "active") -> str:
    """Нормализует status к нижнему регистру. NetBox принимает только lowercase."""
    s = str_or_none(v)
    return s.lower() if s else default


def int_or_none(v) -> Optional[int]:
    try:
        return int(v)
    except (TypeError, ValueError):
        return None


def bool_val(v) -> bool:
    if isinstance(v, bool):
        return v
    return str(v).strip().lower() in ("true", "1", "yes", "да")


_OOB_PATTERNS  = {"ipmi", "idrac", "ilo", "ibmc", "bmc"}
_MGMT_PATTERNS = {"management", "mgmt", "ssh"}


def _is_oob(iface_name: str) -> bool:
    n = (iface_name or "").lower()
    return any(p in n for p in _OOB_PATTERNS)


def _is_mgmt(iface_name: str) -> bool:
    n = (iface_name or "").lower()
    return any(p in n for p in _MGMT_PATTERNS)


def read_sheet(wb, sheet_name: str) -> tuple[list[dict], list]:
    """
    Читает лист Excel.
    Возвращает (col_map, rows):
        col_map — {имя_столбца: номер_столбца (1-based)}
        rows    — список объектов строк openpyxl
    Пропускает строки с голубым фоном (примеры) и полностью пустые.
    """
    EXAMPLE_COLOR = "EBF3FB"

    if sheet_name not in wb.sheetnames:
        log.warning(f"Лист '{sheet_name}' не найден — пропускаем.")
        return {}, []

    ws = wb[sheet_name]
    rows_iter = list(ws.iter_rows())
    if not rows_iter:
        return {}, []

    # Первая строка — заголовки
    header_row = rows_iter[0]
    col_map = {}
    for cell in header_row:
        if cell.value:
            col_map[str(cell.value).strip()] = cell.column

    data_rows = []
    for row in rows_iter[1:]:
        # Пропускаем строки-примеры (голубой фон)
        first_cell = row[0]
        if first_cell.fill and first_cell.fill.fgColor:
            color = first_cell.fill.fgColor.rgb
            if color and color.upper().endswith(EXAMPLE_COLOR):
                continue
        # Пропускаем полностью пустые строки
        if all(c.value is None for c in row):
            continue
        data_rows.append(row)

    return col_map, data_rows


# ─── NetBox helpers ───────────────────────────────────────────────────────────

class NetBoxImporter:
    def __init__(self, url: str, token: str, dry_run: bool = False):
        self.nb = pynetbox.api(url, token=token)
        self.dry_run = dry_run
        self._cache: dict[str, dict] = {}  # кэш уже созданных/найденных объектов

        # Проверяем соединение
        try:
            status = self.nb.status()
            log.info(f"Подключён к NetBox {status.get('netbox-version', '?')} — {url}")
        except Exception as e:
            log.error(f"Не удалось подключиться к NetBox: {e}")
            sys.exit(1)

    # ── Универсальный get-or-create ──────────────────────────────────────────

    def get_or_create(self, endpoint, lookup: dict, create_data: dict, label: str):
        """
        Ищет объект по lookup-полям.
        Если не найден — создаёт с данными create_data.
        Возвращает объект NetBox.
        """
        cache_key = f"{label}:{sorted(lookup.items())}"
        if cache_key in self._cache:
            return self._cache[cache_key]

        try:
            obj = endpoint.get(**lookup)
        except pynetbox.RequestError as e:
            log.error(f"Ошибка поиска {label} {lookup}: {e}")
            return None

        if obj:
            log.debug(f"  Найден  {label}: {obj}")
        else:
            if self.dry_run:
                log.info(f"  [DRY-RUN] Создал бы {label}: {create_data}")
                return None
            try:
                obj = endpoint.create(**create_data)
                log.info(f"  Создан  {label}: {obj}")
            except pynetbox.RequestError as e:
                log.error(f"  Ошибка создания {label} {create_data}: {e}")
                return None

        self._cache[cache_key] = obj
        return obj

    # ── Иерархия места ───────────────────────────────────────────────────────

    def get_or_create_region(self, name: str):
        return self.get_or_create(
            self.nb.dcim.regions,
            {"name": name},
            {"name": name, "slug": name.lower().replace(" ", "-")},
            "Region",
        )

    def get_or_create_site(self, name: str, region_name: str):
        region = self.get_or_create_region(region_name)
        return self.get_or_create(
            self.nb.dcim.sites,
            {"name": name},
            {
                "name": name,
                "slug": name.lower().replace(" ", "-").replace("_", "-"),
                "region": region.id if region else None,
                "status": "active",
            },
            "Site",
        )

    def get_or_create_location(self, name: str, site_name: str, region_name: str):
        site = self.get_or_create_site(site_name, region_name)
        return self.get_or_create(
            self.nb.dcim.locations,
            {"name": name, "site_id": site.id if site else None},
            {
                "name": name,
                "slug": name.lower().replace(" ", "-"),
                "site": site.id if site else None,
            },
            "Location",
        )

    def get_or_create_rack(self, name: str, site_name: str, region_name: str,
                           location_name: Optional[str] = None):
        site = self.get_or_create_site(site_name, region_name)
        location = None
        if location_name:
            location = self.get_or_create_location(location_name, site_name, region_name)

        return self.get_or_create(
            self.nb.dcim.racks,
            {"name": name, "site_id": site.id if site else None},
            {
                "name": name,
                "site": site.id if site else None,
                "location": location.id if location else None,
                "status": "active",
            },
            "Rack",
        )

    # ── Устройство: справочники ───────────────────────────────────────────────

    def get_or_create_manufacturer(self, name: str):
        return self.get_or_create(
            self.nb.dcim.manufacturers,
            {"name": name},
            {"name": name, "slug": name.lower().replace(" ", "-")},
            "Manufacturer",
        )

    def get_or_create_device_type(self, model: str, manufacturer_name: str, height: Optional[int] = 1):
        mfr = self.get_or_create_manufacturer(manufacturer_name)
        return self.get_or_create(
            self.nb.dcim.device_types,
            {"model": model, "manufacturer_id": mfr.id if mfr else None},
            {
                "model": model,
                "slug": model.lower().replace(" ", "-").replace("/", "-"),
                "manufacturer": mfr.id if mfr else None,
                "u_height": height or 1,
            },
            "DeviceType",
        )

    def get_or_create_device_role(self, name: str):
        return self.get_or_create(
            self.nb.dcim.device_roles,
            {"name": name},
            {
                "name": name,
                "slug": name.lower().replace(" ", "-"),
                "color": "2196f3",
                "vm_role": False,
            },
            "DeviceRole",
        )

    def get_or_create_platform(self, name: str):
        return self.get_or_create(
            self.nb.dcim.platforms,
            {"name": name},
            {"name": name, "slug": name.lower().replace(" ", "-")},
            "Platform",
        )

    def get_or_create_tenant(self, name: str):
        slug = name.lower().replace(" ", "-")
        # Сначала ищем по имени (case-sensitive в NetBox), потом по slug
        cache_key = f"Tenant:{name.lower()}"
        if cache_key in self._cache:
            return self._cache[cache_key]
        try:
            obj = self.nb.tenancy.tenants.get(name=name) or self.nb.tenancy.tenants.get(slug=slug)
        except pynetbox.RequestError:
            obj = None
        if obj:
            log.debug(f"  Найден  Tenant: {obj}")
            self._cache[cache_key] = obj
            return obj
        if self.dry_run:
            log.info(f"  [DRY-RUN] Создал бы Tenant: {name}")
            return None
        try:
            obj = self.nb.tenancy.tenants.create(name=name, slug=slug)
            log.info(f"  Создан  Tenant: {obj}")
            self._cache[cache_key] = obj
            return obj
        except pynetbox.RequestError as e:
            # Slug занят — значит tenant с таким slug уже есть, ищем по slug
            if "slug" in str(e).lower():
                try:
                    obj = self.nb.tenancy.tenants.get(slug=slug)
                    if obj:
                        log.debug(f"  Найден  Tenant по slug: {obj}")
                        self._cache[cache_key] = obj
                        return obj
                except Exception:
                    pass
            log.error(f"  Ошибка создания Tenant {name}: {e}")
            return None

    # ── Устройство ────────────────────────────────────────────────────────────

    def get_or_create_device(self, row_data: dict):
        name        = row_data["DeviceName"]
        site_name   = row_data["Site"]
        region_name = row_data["Region"]
        rack_name   = row_data.get("Rack")
        location_name = row_data.get("Location")

        site       = self.get_or_create_site(site_name, region_name)
        rack       = self.get_or_create_rack(rack_name, site_name, region_name, location_name) if rack_name else None
        dev_type   = self.get_or_create_device_type(
            row_data["DeviceType"], row_data["Manufacturer"],
            int_or_none(row_data.get("Height"))
        )
        role       = self.get_or_create_device_role(row_data["Role"])
        platform   = self.get_or_create_platform(row_data["Platform"])
        tenant     = self.get_or_create_tenant(row_data["Tenant"]) if row_data.get("Tenant") else None

        face     = row_data.get("RackFace", "front")
        position = int_or_none(row_data.get("Position"))
        status   = (row_data.get("Status") or "active").lower()

        create_data = {
            "name":        name,
            "device_type": dev_type.id if dev_type else None,
            "role":        role.id if role else None,
            "platform":    platform.id if platform else None,
            "site":        site.id if site else None,
            "rack":        rack.id if rack else None,
            "face":        face,
            "position":    position,
            "tenant":      tenant.id if tenant else None,
            "status":      status,
        }

        # Проверяем — устройство уже существует?
        try:
            existing = self.nb.dcim.devices.get(name=name)
        except pynetbox.RequestError:
            existing = None

        if existing:
            log.debug(f"  Найден  Device: {existing}")
            self._cache[f"Device:{sorted({'name': name}.items())}"] = existing
            return existing

        if self.dry_run:
            log.info(f"  [DRY-RUN] Создал бы Device: {create_data}")
            return None

        try:
            obj = self.nb.dcim.devices.create(**create_data)
            log.info(f"  Создан  Device: {obj}")
            return obj
        except pynetbox.RequestError as e:
            err_str = str(e)
            # Позиция занята — пробуем без rack/position
            if "occupied" in err_str or "position" in err_str.lower():
                log.warning(
                    f"  Позиция U{position} в стойке занята — создаю без rack/position. "
                    f"Исправьте Position в Excel или освободите место в NetBox."
                )
                create_data.pop("rack", None)
                create_data.pop("face", None)
                create_data.pop("position", None)
                try:
                    obj = self.nb.dcim.devices.create(**create_data)
                    log.info(f"  Создан  Device (без стойки): {obj}")
                    return obj
                except pynetbox.RequestError as e2:
                    log.error(f"  Ошибка создания Device: {e2}")
                    return None
            log.error(f"  Ошибка создания Device: {e}")
            return None

    # ── VLAN ──────────────────────────────────────────────────────────────────

    def get_or_create_vlan(self, vid: int, name: str, site_name: str,
                           region_name: Optional[str] = None,
                           group_name: Optional[str] = None,
                           status: str = "active",
                           tenant_name: Optional[str] = None):
        # Ищем сайт только если передан region
        site = None
        if site_name and region_name:
            site = self.get_or_create_site(site_name, region_name)
        elif site_name:
            try:
                site = self.nb.dcim.sites.get(name=site_name)
            except Exception:
                pass

        tenant = self.get_or_create_tenant(tenant_name) if tenant_name else None

        lookup = {"vid": vid}
        if site:
            lookup["site_id"] = site.id

        create_data = {
            "vid":    vid,
            "name":   name,
            "site":   site.id if site else None,
            "status": status_val(status),
            "tenant": tenant.id if tenant else None,
        }

        return self.get_or_create(self.nb.ipam.vlans, lookup, create_data, f"VLAN {vid}")

    # ── Interface ─────────────────────────────────────────────────────────────

    def get_or_create_interface(self, device_name: str, iface_name: str,
                                iface_type: str, mode: str,
                                untagged_vid: Optional[int] = None,
                                tagged_vids: Optional[list] = None,
                                description: str = "",
                                enabled: bool = True,
                                mtu: Optional[int] = None):
        try:
            device = self.nb.dcim.devices.get(name=device_name)
        except Exception:
            device = None

        if not device:
            log.error(f"  Устройство '{device_name}' не найдено — пропускаем интерфейс {iface_name}")
            return None

        # Resolve untagged VLAN
        untagged_obj = None
        if untagged_vid:
            try:
                untagged_obj = self.nb.ipam.vlans.get(vid=untagged_vid)
            except Exception:
                pass

        # Resolve tagged VLANs
        tagged_objs = []
        if tagged_vids:
            for vid in tagged_vids:
                try:
                    v = self.nb.ipam.vlans.get(vid=vid)
                    if v:
                        tagged_objs.append(v.id)
                except Exception:
                    pass

        # NetBox mode: access → "access", tagged → "tagged", tagged-all → "tagged-all"
        # Если mode не задан — не передаём поле вообще (NetBox оставит пустым)
        nb_mode = mode.lower() if mode and mode.strip() else None

        create_data = {
            "device":         device.id,
            "name":           iface_name,
            "type":           normalize_iface_type(iface_type or "1000base-t"),
            "mode":           nb_mode,
            "untagged_vlan":  untagged_obj.id if untagged_obj else None,
            "tagged_vlans":   tagged_objs,
            "description":    description or "",
            "enabled":        enabled,
            "mtu":            mtu,
        }

        cache_key = f"Interface:{device_name}:{iface_name}"
        if cache_key in self._cache:
            return self._cache[cache_key]

        try:
            existing = self.nb.dcim.interfaces.get(device_id=device.id, name=iface_name)
        except pynetbox.RequestError as e:
            log.error(f"  Ошибка поиска интерфейса {iface_name}: {e}")
            return None

        if existing:
            # Обновляем если интерфейс уже есть
            if not self.dry_run:
                try:
                    existing.update(create_data)
                    log.info(f"  Обновлён Interface: {device_name} / {iface_name}")
                except pynetbox.RequestError as e:
                    log.error(f"  Ошибка обновления интерфейса {iface_name}: {e}")
            else:
                log.info(f"  [DRY-RUN] Обновил бы Interface: {device_name} / {iface_name}")
            self._cache[cache_key] = existing
            return existing
        else:
            if self.dry_run:
                log.info(f"  [DRY-RUN] Создал бы Interface: {device_name} / {iface_name}")
                return None
            try:
                obj = self.nb.dcim.interfaces.create(**create_data)
                log.info(f"  Создан  Interface: {device_name} / {iface_name}")
                self._cache[cache_key] = obj
                return obj
            except pynetbox.RequestError as e:
                log.error(f"  Ошибка создания интерфейса {iface_name}: {e}")
                return None

    # ── Prefix ────────────────────────────────────────────────────────────────

    def get_or_create_prefix(self, ip: str, prefix_len: int, site_name: str = ""):
        import ipaddress as _ipaddress
        network = str(_ipaddress.ip_interface(f"{ip}/{prefix_len}").network)

        site = None
        if site_name:
            try:
                site = self.nb.dcim.sites.get(name=site_name)
            except Exception:
                pass

        lookup = {"prefix": network}
        create_data = {
            "prefix": network,
            "status": "active",
            "site":   site.id if site else None,
        }
        return self.get_or_create(self.nb.ipam.prefixes, lookup, create_data, f"Prefix {network}")

    # ── IP Address ────────────────────────────────────────────────────────────

    def create_ip_address(self, device_name: str, iface_name: str,
                          ip: str, prefix_len: int,
                          status: str = "active",
                          dns_name: str = "",
                          description: str = "",
                          is_primary: bool = False):
        address = f"{ip}/{prefix_len}"

        try:
            device = self.nb.dcim.devices.get(name=device_name)
        except Exception:
            device = None

        iface = None
        if device:
            try:
                iface = self.nb.dcim.interfaces.get(device_id=device.id, name=iface_name)
            except Exception:
                pass

        create_data = {
            "address":            address,
            "status":             status_val(status),
            "dns_name":           dns_name or "",
            "description":        description or "",
            "assigned_object_type": "dcim.interface" if iface else None,
            "assigned_object_id":   iface.id if iface else None,
        }

        # Создаём префикс если не существует
        self.get_or_create_prefix(ip, prefix_len)

        # Проверяем — вдруг уже существует
        try:
            existing = self.nb.ipam.ip_addresses.get(address=address)
        except Exception:
            existing = None

        if existing:
            log.info(f"  Уже есть IP: {address} (id={existing.id})")
            ip_obj = existing
        else:
            if self.dry_run:
                log.info(f"  [DRY-RUN] Создал бы IP: {address}")
                return
            try:
                ip_obj = self.nb.ipam.ip_addresses.create(**create_data)
                log.info(f"  Создан  IP: {address} → {device_name}/{iface_name}")
            except pynetbox.RequestError as e:
                log.error(f"  Ошибка создания IP {address}: {e}")
                return

        # Назначаем IP по типу интерфейса (или по флагу IsPrimary из Excel)
        if device and ip_obj and not self.dry_run:
            try:
                if _is_oob(iface_name):
                    # IPMI / iDRAC / iLO / iBMC / BMC → OOB IP
                    device.update({"oob_ip": ip_obj.id})
                    log.info(f"  Назначен OOB IP {address} для {device_name} (интерфейс: {iface_name})")
                elif is_primary or _is_mgmt(iface_name):
                    # Management / SSH / явный флаг IsPrimary → Primary IP
                    if ":" in ip:
                        device.update({"primary_ip6": ip_obj.id})
                    else:
                        device.update({"primary_ip4": ip_obj.id})
                    log.info(f"  Назначен primary IP {address} для {device_name} (интерфейс: {iface_name})")
            except Exception as e:
                log.error(f"  Ошибка назначения IP {address}: {e}")


# ─── Импорт по листам ────────────────────────────────────────────────────────

def import_devices(nb: NetBoxImporter, wb, col_map, rows):
    log.info("─── Импорт устройств ───")
    for row in rows:
        cv = lambda k: cell_val(row, col_map, k)
        device_name = str_or_none(cv("DeviceName"))
        if not device_name:
            continue

        row_data = {
            "DeviceName":   device_name,
            "Region":       str_or_none(cv("Region")) or "Default",
            "Site":         str_or_none(cv("Site")) or "Default",
            "Rack":         str_or_none(cv("Rack")),
            "RackFace":     str_or_none(cv("RackFace")) or "front",
            "Location":     str_or_none(cv("Location")),
            "Role":         str_or_none(cv("Role")) or "Network Device",
            "Manufacturer": str_or_none(cv("Manufacturer")) or "Generic",
            "DeviceType":   str_or_none(cv("DeviceType")) or "Unknown",
            "Height":       cv("Height"),
            "Platform":     str_or_none(cv("Platform")) or "eos",
            "Tenant":       str_or_none(cv("Tenant")),
            "Position":     cv("Position"),
            "Status":       str_or_none(cv("Status")) or "active",
            "Comments":     str_or_none(cv("Comments")),
        }

        log.info(f"Устройство: {device_name}")
        nb.get_or_create_device(row_data)


def import_vlans(nb: NetBoxImporter, wb, col_map, rows):
    log.info("─── Импорт VLANs ───")
    # Нужен регион для поиска сайта — попробуем взять из кэша устройств
    for row in rows:
        cv = lambda k: cell_val(row, col_map, k)
        site_name = str_or_none(cv("Site"))
        vid       = int_or_none(cv("VID"))
        name      = str_or_none(cv("Name"))
        if not vid or not name:
            log.warning(f"  Пропускаем строку: VID={vid} Name={name}")
            continue

        log.info(f"VLAN: {vid} {name} (site={site_name})")
        nb.get_or_create_vlan(
            vid=vid,
            name=name,
            site_name=site_name or "",
            status=status_val(cv("Status")),
            tenant_name=str_or_none(cv("Tenant")),
        )


def import_interfaces(nb: NetBoxImporter, wb, col_map, rows):
    log.info("─── Импорт интерфейсов ───")
    for row in rows:
        cv = lambda k: cell_val(row, col_map, k)
        device_name = str_or_none(cv("DeviceName"))
        iface_name  = str_or_none(cv("InterfaceName"))
        if not device_name or not iface_name:
            continue

        # Парсим TaggedVLANs: "10,20,30,100" → [10, 20, 30, 100]
        tagged_raw = str_or_none(cv("TaggedVLANs"))
        tagged_vids = []
        if tagged_raw:
            for part in tagged_raw.split(","):
                v = int_or_none(part.strip())
                if v:
                    tagged_vids.append(v)

        log.info(f"Interface: {device_name} / {iface_name}")
        nb.get_or_create_interface(
            device_name  = device_name,
            iface_name   = iface_name,
            iface_type   = str_or_none(cv("Type")) or "1000base-t",
            mode         = str_or_none(cv("Mode")),
            untagged_vid = int_or_none(cv("UntaggedVLAN")),
            tagged_vids  = tagged_vids,
            description  = str_or_none(cv("Description")) or "",
            enabled      = bool_val(cv("Enabled") if cv("Enabled") is not None else True),
            mtu          = int_or_none(cv("MTU")),
        )


def import_ip_addresses(nb: NetBoxImporter, wb, col_map, rows):
    log.info("─── Импорт IP-адресов ───")
    for row in rows:
        cv = lambda k: cell_val(row, col_map, k)
        device_name = str_or_none(cv("DeviceName"))
        iface_name  = str_or_none(cv("InterfaceName"))
        ip          = str_or_none(cv("IPAddress"))
        prefix_len  = int_or_none(cv("PrefixLength"))

        if not ip or not prefix_len:
            continue

        log.info(f"IP: {ip}/{prefix_len} → {device_name}/{iface_name}")
        nb.create_ip_address(
            device_name  = device_name or "",
            iface_name   = iface_name or "",
            ip           = ip,
            prefix_len   = prefix_len,
            status       = status_val(cv("Status")),
            dns_name     = str_or_none(cv("DNSName")) or "",
            description  = str_or_none(cv("Description")) or "",
            is_primary   = bool_val(cv("IsPrimary") if cv("IsPrimary") is not None else False),
        )


# ─── Entry point ─────────────────────────────────────────────────────────────

SHEET_IMPORTERS = {
    "Devices":     import_devices,
    "VLANs":       import_vlans,
    "Interfaces":  import_interfaces,
    "IPAddresses": import_ip_addresses,
}

# Порядок важен — устройства до интерфейсов, вланы до интерфейсов
IMPORT_ORDER = ["Devices", "VLANs", "Interfaces", "IPAddresses"]


def main():
    parser = argparse.ArgumentParser(description="Импорт из Excel в NetBox")
    parser.add_argument("--file",    required=True, help="Путь к Excel-файлу")
    parser.add_argument("--dry-run", action="store_true", help="Не записывать в NetBox, только показать что будет")
    parser.add_argument("--sheet",   help="Импортировать только один лист (Devices/VLANs/Interfaces/IPAddresses)")
    parser.add_argument("--verbose", action="store_true", help="Подробный вывод")
    args = parser.parse_args()

    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)

    # NetBox credentials
    netbox_url   = os.environ.get("NETBOX_URL", "http://localhost:8000")
    netbox_token = os.environ.get("NETBOX_TOKEN", "")
    if not netbox_token:
        log.error("NETBOX_TOKEN не задан. Создайте файл .env или задайте переменную окружения.")
        sys.exit(1)

    if args.dry_run:
        log.info("⚠️  DRY-RUN режим — изменений в NetBox не будет")

    # Открываем Excel
    if not os.path.exists(args.file):
        log.error(f"Файл не найден: {args.file}")
        sys.exit(1)

    log.info(f"Читаем файл: {args.file}")
    wb = load_workbook(args.file, data_only=True)

    # Создаём импортёр
    nb = NetBoxImporter(netbox_url, netbox_token, dry_run=args.dry_run)

    # Определяем какие листы обрабатываем
    sheets_to_run = [args.sheet] if args.sheet else IMPORT_ORDER

    for sheet_name in sheets_to_run:
        if sheet_name not in SHEET_IMPORTERS:
            log.warning(f"Неизвестный лист: {sheet_name}. Доступны: {list(SHEET_IMPORTERS.keys())}")
            continue

        col_map, rows = read_sheet(wb, sheet_name)
        if not rows:
            log.info(f"Лист '{sheet_name}': данных нет, пропускаем.")
            continue

        log.info(f"Лист '{sheet_name}': {len(rows)} строк(и)")
        SHEET_IMPORTERS[sheet_name](nb, wb, col_map, rows)

    log.info("✓ Импорт завершён")


if __name__ == "__main__":
    main()
