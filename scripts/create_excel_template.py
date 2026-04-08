"""
Генератор шаблона netbox_import.xlsx.

Использование:
    pip install openpyxl
    python create_excel_template.py
    # создаст netbox_import.xlsx рядом со скриптом
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

OUT_FILE = Path(__file__).parent / "netbox_import.xlsx"

# Цвета
COLOR_HEADER  = "1F4E79"   # тёмно-синий — заголовки
COLOR_EXAMPLE = "EBF3FB"   # голубой — строки-примеры (пропускаются импортом)
COLOR_SHEET   = "D6E4F0"   # подсветка вкладок

# Максимальное количество строк с данными (для named range и dropdown)
MAX_ROWS = 500


# ─── Стили ────────────────────────────────────────────────────────────────────

def header_font():
    return Font(name="Calibri", bold=True, color="FFFFFF", size=11)

def header_fill():
    return PatternFill("solid", fgColor=COLOR_HEADER)

def example_fill():
    return PatternFill("solid", fgColor=COLOR_EXAMPLE)

def thin_border():
    side = Side(style="thin", color="AAAAAA")
    return Border(left=side, right=side, top=side, bottom=side)

def apply_header(ws, columns):
    """Записывает строку заголовков и форматирует её."""
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font    = header_font()
        cell.fill    = header_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border  = thin_border()
        ws.row_dimensions[1].height = 30

def apply_example_row(ws, columns, values):
    """Записывает строку-пример с голубым фоном (импортёр её пропускает)."""
    fill = example_fill()
    italic = Font(name="Calibri", italic=True, color="555555", size=10)
    for col_idx, val in enumerate(values, start=1):
        if col_idx > len(columns):
            break
        cell = ws.cell(row=2, column=col_idx, value=val)
        cell.fill      = fill
        cell.font      = italic
        cell.alignment = Alignment(horizontal="left")
        cell.border    = thin_border()

def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 40)

def freeze_header(ws):
    ws.freeze_panes = "A2"


# ─── Листы ────────────────────────────────────────────────────────────────────

def create_devices_sheet(wb):
    ws = wb.create_sheet("Devices")
    ws.sheet_properties.tabColor = "1F4E79"

    columns = [
        "DeviceName", "Region", "Site", "Location", "Rack",
        "Manufacturer", "DeviceType", "Height", "Position", "RackFace",
        "Role", "Platform", "Status", "Serial", "Tenant",
    ]
    apply_header(ws, columns)
    apply_example_row(ws, columns, [
        "sw-access-01", "HQ", "Moscow-DC", "Floor-1", "Rack-A1",
        "Arista", "DCS-7050CX3-32S", 1, 10, "front",
        "access", "eos", "active", "ABC123456", "",
    ])

    # Dropdown для Platform
    dv_platform = DataValidation(
        type="list",
        formula1='"eos,vrp,comware,ios,nxos,junos"',
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Неверное значение",
        error="Выберите платформу из списка",
    )
    # Dropdown для Status
    dv_status = DataValidation(
        type="list",
        formula1='"active,planned,staged,failed,inventory,decommissioning,offline"',
        showDropDown=False,
    )
    # Dropdown для RackFace
    dv_face = DataValidation(
        type="list",
        formula1='"front,rear"',
        showDropDown=False,
    )

    ws.add_data_validation(dv_platform)
    ws.add_data_validation(dv_status)
    ws.add_data_validation(dv_face)

    platform_col = get_column_letter(columns.index("Platform") + 1)
    status_col   = get_column_letter(columns.index("Status") + 1)
    face_col     = get_column_letter(columns.index("RackFace") + 1)

    dv_platform.sqref = f"{platform_col}3:{platform_col}{MAX_ROWS}"
    dv_status.sqref   = f"{status_col}3:{status_col}{MAX_ROWS}"
    dv_face.sqref     = f"{face_col}3:{face_col}{MAX_ROWS}"

    freeze_header(ws)
    auto_width(ws)
    return ws


def create_vrfs_sheet(wb):
    ws = wb.create_sheet("VRFs")
    ws.sheet_properties.tabColor = "7B2D8B"

    columns = ["Name", "RD", "ImportTargets", "ExportTargets", "Description"]
    apply_header(ws, columns)
    apply_example_row(ws, columns, [
        "PROD", "65000:100", "65000:100", "65000:100", "Production VRF",
    ])

    freeze_header(ws)
    auto_width(ws)
    return ws


def create_vlans_sheet(wb):
    ws = wb.create_sheet("VLANs")
    ws.sheet_properties.tabColor = "107C10"

    columns = ["VID", "Name", "Site", "Status", "Tenant"]
    apply_header(ws, columns)
    apply_example_row(ws, columns, [
        10, "Data", "Moscow-DC", "active", "",
    ])

    dv_status = DataValidation(
        type="list",
        formula1='"active,reserved,deprecated"',
        showDropDown=False,
    )
    ws.add_data_validation(dv_status)
    status_col = get_column_letter(columns.index("Status") + 1)
    dv_status.sqref = f"{status_col}3:{status_col}{MAX_ROWS}"

    freeze_header(ws)
    auto_width(ws)
    return ws


def create_interfaces_sheet(wb, device_name_range: str):
    ws = wb.create_sheet("Interfaces")
    ws.sheet_properties.tabColor = "D83B01"

    columns = [
        "DeviceName", "InterfaceName", "Type", "Mode",
        "UntaggedVLAN", "TaggedVLANs", "Description",
        "Enabled", "MTU", "VRF",
    ]
    apply_header(ws, columns)
    apply_example_row(ws, columns, [
        "sw-access-01", "Ethernet1", "1000base-t", "access",
        10, "", "Uplink to core", "True", 9000, "",
    ])

    # Dropdown DeviceName → из листа Devices
    dv_device = DataValidation(
        type="list",
        formula1=device_name_range,
        showDropDown=False,
        showErrorMessage=False,
    )
    # Dropdown Type
    dv_type = DataValidation(
        type="list",
        formula1='"1000base-t,10gbase-x-sfpp,25gbase-x-sfp28,40gbase-x-qsfpp,100gbase-x-qsfp28,virtual,lag"',
        showDropDown=False,
    )
    # Dropdown Mode
    dv_mode = DataValidation(
        type="list",
        formula1='"access,tagged,tagged-all"',
        showDropDown=False,
    )
    # Dropdown Enabled
    dv_enabled = DataValidation(
        type="list",
        formula1='"True,False"',
        showDropDown=False,
    )

    ws.add_data_validation(dv_device)
    ws.add_data_validation(dv_type)
    ws.add_data_validation(dv_mode)
    ws.add_data_validation(dv_enabled)

    device_col  = get_column_letter(columns.index("DeviceName") + 1)
    type_col    = get_column_letter(columns.index("Type") + 1)
    mode_col    = get_column_letter(columns.index("Mode") + 1)
    enabled_col = get_column_letter(columns.index("Enabled") + 1)

    dv_device.sqref  = f"{device_col}3:{device_col}{MAX_ROWS}"
    dv_type.sqref    = f"{type_col}3:{type_col}{MAX_ROWS}"
    dv_mode.sqref    = f"{mode_col}3:{mode_col}{MAX_ROWS}"
    dv_enabled.sqref = f"{enabled_col}3:{enabled_col}{MAX_ROWS}"

    freeze_header(ws)
    auto_width(ws)
    return ws


def create_ip_sheet(wb, device_name_range: str):
    ws = wb.create_sheet("IPAddresses")
    ws.sheet_properties.tabColor = "E81123"

    columns = [
        "DeviceName", "InterfaceName", "IPAddress", "PrefixLength",
        "Status", "DNSName", "Description", "IsPrimary",
    ]
    apply_header(ws, columns)
    apply_example_row(ws, columns, [
        "sw-access-01", "Management0", "10.0.0.1", 24,
        "active", "sw-access-01.example.com", "Management IP", "True",
    ])

    # Dropdown DeviceName → из листа Devices
    dv_device = DataValidation(
        type="list",
        formula1=device_name_range,
        showDropDown=False,
        showErrorMessage=False,
    )
    dv_status = DataValidation(
        type="list",
        formula1='"active,reserved,deprecated,dhcp,slaac"',
        showDropDown=False,
    )
    dv_primary = DataValidation(
        type="list",
        formula1='"True,False"',
        showDropDown=False,
    )

    ws.add_data_validation(dv_device)
    ws.add_data_validation(dv_status)
    ws.add_data_validation(dv_primary)

    device_col  = get_column_letter(columns.index("DeviceName") + 1)
    status_col  = get_column_letter(columns.index("Status") + 1)
    primary_col = get_column_letter(columns.index("IsPrimary") + 1)

    dv_device.sqref  = f"{device_col}3:{device_col}{MAX_ROWS}"
    dv_status.sqref  = f"{status_col}3:{status_col}{MAX_ROWS}"
    dv_primary.sqref = f"{primary_col}3:{primary_col}{MAX_ROWS}"

    freeze_header(ws)
    auto_width(ws)
    return ws


# ─── Точка входа ──────────────────────────────────────────────────────────────

def main():
    wb = Workbook()
    wb.remove(wb.active)  # удаляем дефолтный лист Sheet

    # Named range для DeviceName: Devices!$A$3:$A$502
    # Используем строку-формулу для data validation (Excel требует кавычки вокруг имени листа)
    device_name_range = f"Devices!$A$3:$A${MAX_ROWS}"

    create_devices_sheet(wb)
    create_vrfs_sheet(wb)
    create_vlans_sheet(wb)
    create_interfaces_sheet(wb, device_name_range)
    create_ip_sheet(wb, device_name_range)

    wb.save(OUT_FILE)
    print(f"Создан файл: {OUT_FILE}")
    print("Заполните лист Devices → DeviceName появится в dropdown на листах Interfaces и IPAddresses")


if __name__ == "__main__":
    main()
