"""
NetBox Custom Script — импорт из Excel.

Установка:
  Скопировать этот файл в директорию scripts/ вашего NetBox.
  NetBox UI → Customization → Scripts → появится "Excel Import"

Зависимость: openpyxl должен быть установлен в окружении NetBox.
  docker exec -it netbox pip install openpyxl   (для Docker)
"""

import re
import ipaddress as _ipaddress
from io import BytesIO

from extras.scripts import Script, FileVar, BooleanVar
from django.contrib.contenttypes.models import ContentType

from dcim.models import (
    Region, Site, Location, Rack,
    Manufacturer, DeviceType, DeviceRole, Platform,
    Device, Interface,
)
from ipam.models import VLAN, IPAddress, Prefix
from tenancy.models import Tenant
from extras.models import Tag

from openpyxl import load_workbook


# ─── Утилиты ─────────────────────────────────────────────────────────────────

EXAMPLE_COLOR = "EBF3FB"

IFACE_TYPE_MAP = {
    "1000base-t":      "1000base-t",
    "10gbase-t":       "10gbase-t",
    "1g":              "1000base-t",
    "10g":             "10gbase-x-sfpp",
    "sfp+":            "10gbase-x-sfpp",
    "25g":             "25gbase-x-sfp28",
    "40g":             "40gbase-x-qsfpp",
    "qsfp+":           "40gbase-x-qsfpp",
    "100g":            "100gbase-x-qsfp28",
    "100gbase-t":      "100gbase-x-qsfp28",
    "100gbase-sr4":    "100gbase-x-qsfp28",
    "100gbase-lr4":    "100gbase-x-qsfp28",
    "qsfp28":          "100gbase-x-qsfp28",
    "400g":            "400gbase-x-qsfpdd",
    "virtual":         "virtual",
    "lag":             "lag",
    "port-channel":    "lag",
}


def slugify(s: str) -> str:
    s = s.lower().strip()
    s = re.sub(r"[^\w\s-]", "", s)
    s = re.sub(r"[\s_]+", "-", s)
    s = re.sub(r"-+", "-", s)
    return s[:100]


def normalize_type(t: str) -> str:
    return IFACE_TYPE_MAP.get((t or "").strip().lower(), t or "1000base-t")


def normalize_status(v, default="active") -> str:
    return (v or default).strip().lower()


def str_val(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def int_val(v) -> int | None:
    try:
        return int(v)
    except (TypeError, ValueError):
        return None


def bool_val(v) -> bool:
    if isinstance(v, bool):
        return v
    return str(v).strip().lower() in ("true", "1", "yes")


def read_sheet(wb, sheet_name: str):
    """
    Возвращает (col_map, rows).
    Пропускает строки-примеры (голубой фон) и пустые строки.
    """
    if sheet_name not in wb.sheetnames:
        return {}, []

    ws = wb[sheet_name]
    all_rows = list(ws.iter_rows())
    if not all_rows:
        return {}, []

    col_map = {
        str(cell.value).strip(): cell.column
        for cell in all_rows[0]
        if cell.value
    }

    data_rows = []
    for row in all_rows[1:]:
        # Пропускаем строки-примеры
        fc = row[0]
        if fc.fill and fc.fill.fgColor and fc.fill.fgColor.rgb:
            if fc.fill.fgColor.rgb.upper().endswith(EXAMPLE_COLOR):
                continue
        # Пропускаем пустые
        if all(c.value is None for c in row):
            continue
        data_rows.append(row)

    return col_map, data_rows


def cv(row, col_map, key):
    idx = col_map.get(key)
    if idx is None:
        return None
    val = row[idx - 1].value
    if val is None:
        return None
    return str(val).strip() if isinstance(val, str) else val


# Паттерны для автоопределения типа IP по имени интерфейса
_OOB_PATTERNS  = {"ipmi", "idrac", "ilo", "ibmc", "bmc"}
_MGMT_PATTERNS = {"management", "mgmt", "ssh"}


def _is_oob(iface_name: str) -> bool:
    n = (iface_name or "").lower()
    return any(p in n for p in _OOB_PATTERNS)


def _is_mgmt(iface_name: str) -> bool:
    n = (iface_name or "").lower()
    return any(p in n for p in _MGMT_PATTERNS)


# Платформы сетевого оборудования — получат тег "config-pending"
NETWORK_PLATFORMS = {"eos", "arista-eos", "arista", "vrp", "comware", "ios", "nxos", "junos"}
# Роли серверов — им тег НЕ ставится
SERVER_ROLE_SLUGS = {"server", "servers", "vm", "virtual-machine", "esxi", "hypervisor"}


def _get_or_create_tag(slug: str, name: str, color: str = "ff9800") -> Tag:
    tag, _ = Tag.objects.get_or_create(
        slug=slug,
        defaults={"name": name, "color": color},
    )
    return tag


def _is_network_device(platform_name: str, role_slug: str) -> bool:
    """True если устройство — сетевое оборудование, а не сервер."""
    if role_slug and role_slug.lower() in SERVER_ROLE_SLUGS:
        return False
    if platform_name and platform_name.lower() in NETWORK_PLATFORMS:
        return True
    return False


# ─── Хелперы get-or-create через Django ORM ──────────────────────────────────

def get_or_create_region(name: str):
    obj, created = Region.objects.get_or_create(
        name=name,
        defaults={"slug": slugify(name)},
    )
    return obj, created


def get_or_create_site(name: str, region_name: str):
    region, _ = get_or_create_region(region_name)
    obj, created = Site.objects.get_or_create(
        name=name,
        defaults={"slug": slugify(name), "region": region, "status": "active"},
    )
    return obj, created


def get_or_create_location(name: str, site: Site):
    obj, created = Location.objects.get_or_create(
        name=name,
        site=site,
        defaults={"slug": slugify(name)},
    )
    return obj, created


def get_or_create_rack(name: str, site: Site, location=None):
    obj, created = Rack.objects.get_or_create(
        name=name,
        site=site,
        defaults={"status": "active", "location": location},
    )
    return obj, created


def get_or_create_manufacturer(name: str):
    obj, created = Manufacturer.objects.get_or_create(
        name=name,
        defaults={"slug": slugify(name)},
    )
    return obj, created


def get_or_create_device_type(model: str, manufacturer: Manufacturer, height: int = 1):
    obj, created = DeviceType.objects.get_or_create(
        manufacturer=manufacturer,
        model=model,
        defaults={"slug": slugify(model), "u_height": height or 1},
    )
    return obj, created


def get_or_create_role(name: str):
    obj, created = DeviceRole.objects.get_or_create(
        name=name,
        defaults={"slug": slugify(name), "color": "2196f3"},
    )
    return obj, created


def get_or_create_platform(name: str):
    obj, created = Platform.objects.get_or_create(
        name=name,
        defaults={"slug": slugify(name)},
    )
    return obj, created


def get_or_create_tenant(name: str):
    slug = slugify(name)
    # Ищем сначала по имени, потом по slug (разный регистр)
    obj = Tenant.objects.filter(name__iexact=name).first() \
       or Tenant.objects.filter(slug=slug).first()
    if obj:
        return obj, False
    obj = Tenant.objects.create(name=name, slug=slug)
    return obj, True


# ─── Основной скрипт ─────────────────────────────────────────────────────────

class ExcelImport(Script):

    class Meta:
        name = "Excel Import"
        description = "Импорт устройств, VLANs, интерфейсов и IP из Excel-файла"
        field_order = ["excel_file", "dry_run"]

    excel_file = FileVar(
        label="Excel файл",
        description="Файл формата netbox_import.xlsx (листы: Devices, VLANs, Interfaces, IPAddresses)",
    )
    dry_run = BooleanVar(
        label="Dry Run",
        description="Показать что будет создано — без записи в базу",
        default=False,
    )

    def run(self, data, commit):
        dry = data["dry_run"]
        if dry:
            self.log_warning("DRY RUN — изменений в базе не будет")

        # Читаем файл из памяти
        file_obj = data["excel_file"]
        content = file_obj.read() if hasattr(file_obj, "read") else file_obj
        wb = load_workbook(BytesIO(content), data_only=True)

        self._import_devices(wb, dry)
        self._import_vlans(wb, dry)
        self._import_interfaces(wb, dry)
        self._import_ips(wb, dry)

        if dry:
            # Откатываем транзакцию — ничего не сохраняется
            raise Exception("DRY RUN завершён — транзакция откатана")

    # ── Devices ───────────────────────────────────────────────────────────────

    def _tag_if_network(self, device: Device, platform_name: str, role_name: str):
        """Ставит тег config-pending если устройство — сетевое оборудование."""
        role_slug = (role_name or "").lower().replace(" ", "-")
        if _is_network_device(platform_name, role_slug):
            tag = _get_or_create_tag("config-pending", "config-pending", "ff9800")
            device.tags.add(tag)
            self.log_success(f"  Тег 'config-pending' добавлен → {device.name}")

    def _import_devices(self, wb, dry: bool):
        col_map, rows = read_sheet(wb, "Devices")
        if not rows:
            self.log_info("Лист Devices: данных нет")
            return

        self.log_info(f"Лист Devices: {len(rows)} строк(и)")

        for row in rows:
            name = str_val(cv(row, col_map, "DeviceName"))
            if not name:
                continue

            region_name = str_val(cv(row, col_map, "Region")) or "Default"
            site_name   = str_val(cv(row, col_map, "Site")) or "Default"
            rack_name   = str_val(cv(row, col_map, "Rack"))
            loc_name    = str_val(cv(row, col_map, "Location"))
            mfr_name    = str_val(cv(row, col_map, "Manufacturer")) or "Generic"
            dtype_name  = str_val(cv(row, col_map, "DeviceType")) or "Unknown"
            role_name   = str_val(cv(row, col_map, "Role")) or "Network Device"
            platform_nm = str_val(cv(row, col_map, "Platform")) or "eos"
            tenant_name = str_val(cv(row, col_map, "Tenant"))
            height      = int_val(cv(row, col_map, "Height")) or 1
            position    = int_val(cv(row, col_map, "Position"))
            face        = str_val(cv(row, col_map, "RackFace")) or "front"
            status      = normalize_status(cv(row, col_map, "Status"))
            serial      = str_val(cv(row, col_map, "Serial")) or ""

            # Уже существует?
            if Device.objects.filter(name=name).exists():
                self.log_info(f"Устройство уже существует: {name}")
                continue

            if dry:
                self.log_success(f"[DRY] Создал бы Device: {name} ({site_name})")
                continue

            # Создаём зависимости
            site, _  = get_or_create_site(site_name, region_name)
            mfr, _   = get_or_create_manufacturer(mfr_name)
            dtype, _ = get_or_create_device_type(dtype_name, mfr, height)
            role, _  = get_or_create_role(role_name)
            platform, _ = get_or_create_platform(platform_nm)
            tenant    = get_or_create_tenant(tenant_name)[0] if tenant_name else None

            location = None
            if loc_name:
                location, _ = get_or_create_location(loc_name, site)

            rack = None
            if rack_name:
                rack, _ = get_or_create_rack(rack_name, site, location)

            try:
                Device.objects.create(
                    name=name,
                    device_type=dtype,
                    role=role,
                    platform=platform,
                    site=site,
                    rack=rack,
                    face=face if rack else "",
                    position=position if rack else None,
                    tenant=tenant,
                    status=status,
                    serial=serial,
                )
                self.log_success(f"Создано устройство: {name}")
                self._tag_if_network(
                    Device.objects.get(name=name),
                    platform_nm,
                    role_name,
                )
            except Exception as e:
                err = str(e)
                if "occupied" in err or "position" in err.lower():
                    self.log_warning(f"Позиция U{position} занята — создаю {name} без стойки")
                    dev = Device.objects.create(
                        name=name,
                        device_type=dtype,
                        role=role,
                        platform=platform,
                        site=site,
                        tenant=tenant,
                        status=status,
                        serial=serial,
                    )
                    self.log_success(f"Создано устройство (без стойки): {name}")
                    self._tag_if_network(dev, platform_nm, role_name)
                else:
                    self.log_failure(f"Ошибка создания {name}: {e}")

    # ── VLANs ─────────────────────────────────────────────────────────────────

    def _import_vlans(self, wb, dry: bool):
        col_map, rows = read_sheet(wb, "VLANs")
        if not rows:
            self.log_info("Лист VLANs: данных нет")
            return

        self.log_info(f"Лист VLANs: {len(rows)} строк(и)")

        for row in rows:
            vid       = int_val(cv(row, col_map, "VID"))
            vlan_name = str_val(cv(row, col_map, "Name"))
            site_name = str_val(cv(row, col_map, "Site"))
            status    = normalize_status(cv(row, col_map, "Status"))
            tenant_nm = str_val(cv(row, col_map, "Tenant"))

            if not vid or not vlan_name:
                continue

            site = Site.objects.filter(name=site_name).first() if site_name else None
            tenant = get_or_create_tenant(tenant_nm)[0] if tenant_nm else None

            if dry:
                self.log_success(f"[DRY] Создал бы VLAN {vid} {vlan_name}")
                continue

            vlan, created = VLAN.objects.get_or_create(
                vid=vid,
                site=site,
                defaults={"name": vlan_name, "status": status, "tenant": tenant},
            )
            if created:
                self.log_success(f"Создан VLAN {vid}: {vlan_name}")
            else:
                self.log_info(f"VLAN уже существует: {vid} {vlan_name}")

    # ── Interfaces ────────────────────────────────────────────────────────────

    def _import_interfaces(self, wb, dry: bool):
        col_map, rows = read_sheet(wb, "Interfaces")
        if not rows:
            self.log_info("Лист Interfaces: данных нет")
            return

        self.log_info(f"Лист Interfaces: {len(rows)} строк(и)")

        for row in rows:
            device_name = str_val(cv(row, col_map, "DeviceName"))
            iface_name  = str_val(cv(row, col_map, "InterfaceName"))
            if not device_name or not iface_name:
                continue

            iface_type   = normalize_type(str_val(cv(row, col_map, "Type")))
            mode         = str_val(cv(row, col_map, "Mode"))
            nb_mode      = mode.lower() if mode and mode.strip() else None
            untagged_vid = int_val(cv(row, col_map, "UntaggedVLAN"))
            tagged_raw   = str_val(cv(row, col_map, "TaggedVLANs"))
            description  = str_val(cv(row, col_map, "Description")) or ""
            enabled      = bool_val(cv(row, col_map, "Enabled")) if cv(row, col_map, "Enabled") is not None else True
            mtu          = int_val(cv(row, col_map, "MTU"))

            tagged_vids = []
            if tagged_raw:
                for part in tagged_raw.split(","):
                    v = int_val(part.strip())
                    if v:
                        tagged_vids.append(v)

            if dry:
                self.log_success(f"[DRY] Создал бы Interface: {device_name} / {iface_name} mode={nb_mode}")
                continue

            device = Device.objects.filter(name=device_name).first()
            if not device:
                self.log_failure(f"Устройство не найдено: {device_name} — пропускаем {iface_name}")
                continue

            untagged_vlan = VLAN.objects.filter(vid=untagged_vid).first() if untagged_vid else None
            tagged_vlans  = list(VLAN.objects.filter(vid__in=tagged_vids)) if tagged_vids else []

            defaults = {
                "type":          iface_type,
                "mode":          nb_mode or "",
                "description":   description,
                "enabled":       enabled,
                "mtu":           mtu,
                "untagged_vlan": untagged_vlan,
            }

            iface, created = Interface.objects.get_or_create(
                device=device,
                name=iface_name,
                defaults=defaults,
            )

            if not created:
                # Обновляем существующий
                for k, v in defaults.items():
                    setattr(iface, k, v)
                iface.save()
                self.log_info(f"Обновлён интерфейс: {device_name} / {iface_name}")
            else:
                self.log_success(f"Создан интерфейс: {device_name} / {iface_name}")

            if tagged_vlans:
                iface.tagged_vlans.set(tagged_vlans)

    # ── IP Addresses ──────────────────────────────────────────────────────────

    def _import_ips(self, wb, dry: bool):
        col_map, rows = read_sheet(wb, "IPAddresses")
        if not rows:
            self.log_info("Лист IPAddresses: данных нет")
            return

        self.log_info(f"Лист IPAddresses: {len(rows)} строк(и)")

        for row in rows:
            device_name = str_val(cv(row, col_map, "DeviceName"))
            iface_name  = str_val(cv(row, col_map, "InterfaceName"))
            ip          = str_val(cv(row, col_map, "IPAddress"))
            prefix_len  = int_val(cv(row, col_map, "PrefixLength"))
            status      = normalize_status(cv(row, col_map, "Status"))
            dns_name    = str_val(cv(row, col_map, "DNSName")) or ""
            description = str_val(cv(row, col_map, "Description")) or ""
            is_primary  = bool_val(cv(row, col_map, "IsPrimary"))

            if not ip or not prefix_len:
                continue

            address = f"{ip}/{prefix_len}"

            if dry:
                network = str(_ipaddress.ip_interface(address).network)
                self.log_success(f"[DRY] Создал бы Prefix: {network}")
                self.log_success(f"[DRY] Создал бы IP: {address} → {device_name}/{iface_name}")
                continue

            # Создаём префикс если не существует
            network = str(_ipaddress.ip_interface(address).network)
            prefix_obj = Prefix.objects.filter(prefix=network, vrf=None).first()
            if prefix_obj is None:
                prefix_obj = Prefix.objects.create(prefix=network, status="active")
                self.log_success(f"Создан Prefix: {network}")
            else:
                self.log_info(f"Prefix уже существует: {network}")

            # Ищем интерфейс
            iface = None
            device = Device.objects.filter(name=device_name).first() if device_name else None
            if device and iface_name:
                iface = Interface.objects.filter(device=device, name=iface_name).first()

            # Тип контента для привязки к интерфейсу
            ct = ContentType.objects.get_for_model(Interface) if iface else None

            ip_obj, created = IPAddress.objects.get_or_create(
                address=address,
                defaults={
                    "status":               status,
                    "dns_name":             dns_name,
                    "description":          description,
                    "assigned_object_type": ct,
                    "assigned_object_id":   iface.pk if iface else None,
                },
            )

            if created:
                self.log_success(f"Создан IP: {address} → {device_name}/{iface_name}")
            else:
                self.log_info(f"IP уже существует: {address}")

            # Назначаем IP по типу интерфейса (или по флагу IsPrimary из Excel)
            if device and ip_obj:
                if _is_oob(iface_name or ""):
                    # IPMI / iDRAC / iLO / iBMC / BMC → OOB IP устройства
                    device.oob_ip = ip_obj
                    device.save()
                    self.log_success(f"Назначен OOB IP {address} для {device_name} (интерфейс: {iface_name})")
                elif is_primary or _is_mgmt(iface_name or "") or _is_mgmt(description or ""):
                    # Management / SSH / явный флаг IsPrimary → Primary IP
                    if ":" in ip:
                        device.primary_ip6 = ip_obj
                    else:
                        device.primary_ip4 = ip_obj
                    device.save()
                    self.log_success(f"Назначен primary IP {address} для {device_name} (интерфейс: {iface_name})")
