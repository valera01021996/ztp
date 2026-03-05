"""
Генератор Excel-шаблонов для импорта данных в NetBox.
Запуск: python create_excel_templates.py
Создаёт файл: ../excel_templates/netbox_import.xlsx
"""

import os
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "excel_templates")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "netbox_import.xlsx")

# ─── Цветовая палитра ───────────────────────────────────────────────────────
CLR_HEADER_DARK   = "1E3A5F"   # тёмно-синий  — заголовок обязательного поля
CLR_HEADER_LIGHT  = "2E6DA4"   # синий         — заголовок опционального поля
CLR_HEADER_TEXT   = "FFFFFF"   # белый текст
CLR_EXAMPLE       = "EBF3FB"   # светло-голубой — строка-пример
CLR_SHEET_TAB = {
    "Instructions": "2ECC71",
    "Devices":      "1E3A5F",
    "VLANs":        "8E44AD",
    "Interfaces":   "E67E22",
    "IPAddresses":  "27AE60",
}

# ─── Вспомогательные функции ────────────────────────────────────────────────

def make_header_cell(ws, row, col, value, optional=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = PatternFill("solid", fgColor=CLR_HEADER_LIGHT if optional else CLR_HEADER_DARK)
    cell.font = Font(color=CLR_HEADER_TEXT, bold=True, size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="AAAAAA")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return cell


def make_example_row(ws, row, values):
    fill = PatternFill("solid", fgColor=CLR_EXAMPLE)
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill = fill
        cell.alignment = Alignment(vertical="center")
        cell.border = border


def set_column_widths(ws, widths):
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width


def add_dropdown(ws, col_letter, start_row, end_row, formula):
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.sqref = f"{col_letter}{start_row}:{col_letter}{end_row}"
    ws.add_data_validation(dv)


def freeze_and_autofilter(ws, freeze_cell, filter_range):
    ws.freeze_panes = freeze_cell
    ws.auto_filter.ref = filter_range


# ─── Лист: Instructions ──────────────────────────────────────────────────────

def build_instructions(wb):
    ws = wb.create_sheet("Instructions", 0)
    ws.sheet_properties.tabColor = CLR_SHEET_TAB["Instructions"]
    ws.column_dimensions["A"].width = 120

    lines = [
        ("ИНСТРУКЦИЯ ПО ЗАПОЛНЕНИЮ NETBOX IMPORT", True, "1E3A5F", "FFFFFF", 16),
        ("", False, None, None, 11),
        ("ЛИСТЫ И ПОРЯДОК ЗАПОЛНЕНИЯ", True, "2E6DA4", "FFFFFF", 13),
        ("  1. Devices      — основные данные об устройствах (обязателен)", False, None, None, 11),
        ("  2. VLANs        — список вланов с именами", False, None, None, 11),
        ("  3. Interfaces   — интерфейсы устройств (аксесс / транк / management)", False, None, None, 11),
        ("  4. IPAddresses  — IP-адреса, привязанные к интерфейсам", False, None, None, 11),
        ("", False, None, None, 11),
        ("ПРАВИЛА ЗАПОЛНЕНИЯ", True, "2E6DA4", "FFFFFF", 13),
        ("  • Строки с примерами (голубой фон) — только для справки, удалите перед импортом", False, None, None, 11),
        ("  • Тёмно-синие заголовки = обязательные поля", False, None, None, 11),
        ("  • Светло-синие заголовки = опциональные поля", False, None, None, 11),
        ("  • Значения чувствительны к регистру там, где указано", False, None, None, 11),
        ("  • Не меняйте названия столбцов и листов — скрипт читает их по имени", False, None, None, 11),
        ("", False, None, None, 11),
        ("ЗАПУСК ИМПОРТА", True, "2E6DA4", "FFFFFF", 13),
        ("  pip install -r requirements.txt", False, None, "1E3A5F", 11),
        ("  cp .env.example .env   # заполните NETBOX_URL и NETBOX_TOKEN", False, None, "1E3A5F", 11),
        ("  python import_to_netbox.py --file ../excel_templates/netbox_import.xlsx", False, None, "1E3A5F", 11),
        ("", False, None, None, 11),
        ("ПОРЯДОК ЗАВИСИМОСТЕЙ (скрипт учитывает автоматически)", True, "2E6DA4", "FFFFFF", 13),
        ("  Region → Site → Location → Rack → Manufacturer → DeviceType → Role → Device", False, None, None, 11),
        ("  Site → VLAN", False, None, None, 11),
        ("  Device → Interface → IP Address", False, None, None, 11),
        ("  VLAN → Interface (untagged/tagged assignment)", False, None, None, 11),
    ]

    for i, (text, bold, bg, fg, size) in enumerate(lines, start=1):
        cell = ws.cell(row=i, column=1, value=text)
        cell.font = Font(bold=bold, size=size, color=fg or "000000",
                         name="Courier New" if bg is None and text.startswith("  python") or text.startswith("  pip") or text.startswith("  cp") else "Calibri")
        if bg:
            cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(vertical="center", indent=0)
        ws.row_dimensions[i].height = 20 if text else 8


# ─── Лист: Devices ───────────────────────────────────────────────────────────

def build_devices(wb):
    ws = wb.create_sheet("Devices")
    ws.sheet_properties.tabColor = CLR_SHEET_TAB["Devices"]

    # Обязательные поля
    required = [
        "Region", "Site", "Rack", "RackFace",
        "Role", "Manufacturer", "DeviceType",
        "Platform", "DeviceName",
    ]
    # Опциональные поля
    optional = [
        "Location", "Height", "Position", "Tenant", "Status", "Serial", "Comments",
    ]

    headers = required + optional
    opt_start = len(required) + 1

    for col, h in enumerate(headers, start=1):
        make_header_cell(ws, 1, col, h, optional=(col >= opt_start))

    # Строка-пример
    make_example_row(ws, 2, [
        "Russia",           # Region
        "MSK-DC1",          # Site
        "RACK-01",          # Rack
        "front",            # RackFace  (front / rear)
        "Access Switch",    # Role
        "Arista",           # Manufacturer
        "DCS-7050CX3-32S",  # DeviceType
        "eos",              # Platform   (eos / vrp / comware)
        "sw-rack01-01",     # DeviceName
        "Server Room A",    # Location   (опционально)
        "2",                # Height (U)
        "10",               # Position   (номер юнита в стойке)
        "ClientA",          # Tenant
        "active",           # Status
        "JPE12345678",      # Serial
        "",                 # Comments
    ])
    make_example_row(ws, 3, [
        "Russia", "MSK-DC1", "RACK-01", "front",
        "Access Switch", "Huawei", "CE6870-48S6CQ",
        "vrp", "sw-rack01-02",
        "Server Room A", "2", "12", "ClientA", "active", "H3C98765432", "",
    ])

    # Dropdown валидации
    add_dropdown(ws, "D", 2, 1000, '"front,rear"')
    add_dropdown(ws, "K", 2, 1000, '"eos,vrp,comware,ios,nxos,junos"')
    add_dropdown(ws, "N", 2, 1000, '"active,planned,staged,failed,offline,decommissioning"')

    set_column_widths(ws, [14, 14, 12, 10, 18, 14, 22, 12, 20, 18, 8, 10, 14, 12, 16, 20])
    freeze_and_autofilter(ws, "A2", f"A1:{get_column_letter(len(headers))}1")
    ws.row_dimensions[1].height = 30


# ─── Лист: VLANs ─────────────────────────────────────────────────────────────

def build_vlans(wb):
    ws = wb.create_sheet("VLANs")
    ws.sheet_properties.tabColor = CLR_SHEET_TAB["VLANs"]

    required = ["Site", "VID", "Name"]
    optional = ["Group", "Status", "Description", "Tenant"]

    headers = required + optional
    opt_start = len(required) + 1

    for col, h in enumerate(headers, start=1):
        make_header_cell(ws, 1, col, h, optional=(col >= opt_start))

    examples = [
        ["MSK-DC1", 10,  "MGMT",        "",  "active", "Management vlan",      ""],
        ["MSK-DC1", 20,  "SERVERS",      "",  "active", "Server traffic",        "ClientA"],
        ["MSK-DC1", 30,  "STORAGE",      "",  "active", "Storage network",       "ClientA"],
        ["MSK-DC1", 100, "UPLINK-TRUNK", "",  "active", "Uplink aggregation",    ""],
        ["MSK-DC1", 200, "VMOTION",      "",  "active", "VMware vMotion",        "ClientA"],
    ]
    for i, row in enumerate(examples, start=2):
        make_example_row(ws, i, row)

    add_dropdown(ws, "E", 2, 1000, '"active,reserved,deprecated"')

    set_column_widths(ws, [14, 8, 20, 16, 12, 30, 14])
    freeze_and_autofilter(ws, "A2", f"A1:{get_column_letter(len(headers))}1")
    ws.row_dimensions[1].height = 30


# ─── Лист: Interfaces ────────────────────────────────────────────────────────

def build_interfaces(wb):
    ws = wb.create_sheet("Interfaces")
    ws.sheet_properties.tabColor = CLR_SHEET_TAB["Interfaces"]

    required = ["DeviceName", "InterfaceName", "Type", "Mode"]
    optional = ["UntaggedVLAN", "TaggedVLANs", "Description", "Enabled", "LAG", "MTU"]

    headers = required + optional
    opt_start = len(required) + 1

    for col, h in enumerate(headers, start=1):
        make_header_cell(ws, 1, col, h, optional=(col >= opt_start))

    examples = [
        # DeviceName,       IfName,         Type,          Mode,        Untag, Tagged,          Desc,                 Enabled, LAG, MTU
        ["sw-rack01-01", "Management0",  "1000base-t",  "access",    10,    "",              "MGMT",               "true",  "",  ""],
        ["sw-rack01-01", "Ethernet1",    "1000base-t",  "access",    20,    "",              "SERVER-01-eth0",      "true",  "",  ""],
        ["sw-rack01-01", "Ethernet2",    "1000base-t",  "access",    20,    "",              "SERVER-01-eth1",      "true",  "",  ""],
        ["sw-rack01-01", "Ethernet3",    "1000base-t",  "access",    30,    "",              "STORAGE-01",          "true",  "",  ""],
        ["sw-rack01-01", "Ethernet48",   "10gbase-t",   "tagged",    "",    "10,20,30,100",  "UPLINK-to-AGG-01",    "true",  "",  "9214"],
        ["sw-rack01-01", "Ethernet49",   "10gbase-t",   "tagged",    "",    "10,20,30,100",  "UPLINK-to-AGG-02",    "true",  "",  "9214"],
        ["sw-rack01-02", "Management0",  "1000base-t",  "access",    10,    "",              "MGMT",               "true",  "",  ""],
        ["sw-rack01-02", "GigabitEthernet0/0/1", "1000base-t", "access", 20, "",            "SERVER-02-eth0",      "true",  "",  ""],
        ["sw-rack01-02", "XGigabitEthernet0/0/1", "10gbase-t", "tagged", "", "10,20,30,100","UPLINK-to-AGG-01",    "true",  "",  "9216"],
    ]
    for i, row in enumerate(examples, start=2):
        make_example_row(ws, i, row)

    # Dropdown для Type
    types = '"1000base-t,10gbase-t,25gbase-x-sfp28,40gbase-x-qsfpp,100gbase-x-qsfp28,virtual,lag"'
    add_dropdown(ws, "C", 2, 1000, types)
    add_dropdown(ws, "D", 2, 1000, '"access,tagged,tagged-all"')
    add_dropdown(ws, "H", 2, 1000, '"true,false"')

    # Комментарий к TaggedVLANs
    from openpyxl.comments import Comment
    comment_cell = ws.cell(row=1, column=6)
    comment = Comment("Перечислите VID через запятую без пробелов.\nПример: 10,20,30,100", "System")
    comment.width = 250
    comment.height = 60
    comment_cell.comment = comment

    set_column_widths(ws, [20, 28, 14, 10, 12, 20, 28, 8, 12, 8])
    freeze_and_autofilter(ws, "A2", f"A1:{get_column_letter(len(headers))}1")
    ws.row_dimensions[1].height = 30


# ─── Лист: IPAddresses ───────────────────────────────────────────────────────

def build_ip_addresses(wb):
    ws = wb.create_sheet("IPAddresses")
    ws.sheet_properties.tabColor = CLR_SHEET_TAB["IPAddresses"]

    required = ["DeviceName", "InterfaceName", "IPAddress", "PrefixLength"]
    optional = ["Status", "DNSName", "Description", "IsPrimary"]

    headers = required + optional
    opt_start = len(required) + 1

    for col, h in enumerate(headers, start=1):
        make_header_cell(ws, 1, col, h, optional=(col >= opt_start))

    examples = [
        ["sw-rack01-01", "Management0", "192.168.10.11", 24, "active", "sw-rack01-01.mgmt.local", "Management IP", "true"],
        ["sw-rack01-02", "Management0", "192.168.10.12", 24, "active", "sw-rack01-02.mgmt.local", "Management IP", "true"],
    ]
    for i, row in enumerate(examples, start=2):
        make_example_row(ws, i, row)

    add_dropdown(ws, "E", 2, 1000, '"active,reserved,deprecated,dhcp,slaac"')
    add_dropdown(ws, "H", 2, 1000, '"true,false"')

    set_column_widths(ws, [20, 28, 18, 14, 12, 28, 28, 10])
    freeze_and_autofilter(ws, "A2", f"A1:{get_column_letter(len(headers))}1")
    ws.row_dimensions[1].height = 30


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    wb = Workbook()
    # Удаляем дефолтный лист
    wb.remove(wb.active)

    build_instructions(wb)
    build_devices(wb)
    build_vlans(wb)
    build_interfaces(wb)
    build_ip_addresses(wb)

    wb.save(OUTPUT_FILE)
    print(f"✓ Excel-шаблон создан: {os.path.abspath(OUTPUT_FILE)}")
    print()
    print("Листы:")
    print("  1. Instructions  — инструкция по заполнению")
    print("  2. Devices       — устройства (обязателен)")
    print("  3. VLANs         — вланы")
    print("  4. Interfaces    — интерфейсы")
    print("  5. IPAddresses   — IP-адреса")


if __name__ == "__main__":
    main()
